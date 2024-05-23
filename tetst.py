import db.database as db
from openpyxl import load_workbook
from telethon import TelegramClient
from telethon.tl.types import InputPhoneContact, InputUser
from telethon.tl.functions.contacts import ImportContactsRequest
from telethon.tl.functions.contacts import DeleteContactsRequest
from telethon.errors.rpcerrorlist import PeerFloodError, UserPrivacyRestrictedError, FloodWaitError
import config_controller
import asyncio

config_controller.preload_config()
async def st():
    is_good = False
    src = "agro.xlsx"
    wb = load_workbook(src)
    sheet = wb["Лист1"]
    name_group = "Агро Полтавська область"
    index = 0

    client = TelegramClient(session="myak",
                            api_id=config_controller.LIST_TG_ACC["myak"]["api_id"],
                            api_hash=config_controller.LIST_TG_ACC["myak"][
                                "api_hash"])
    await client.connect()
    for row in sheet.iter_rows(values_only=True):
        index +=1
        if index < 2:
            continue
        phone = row[12]
        name = row[8]
        if name == "Ноздрін Олександр Олександрович":
            is_good = True
            continue
        if not is_good:
            continue
        if len(phone) != 0:
            print(name, phone, len(phone))
            try:
                print("GET", name, phone)
                contact = InputPhoneContact(client_id=0, phone=phone, first_name=name.split(" ")[0], last_name=name.split(" ")[1])
                result = await client(ImportContactsRequest([contact]))
                tg_id = result.users[0].id
                tg_name = result.users[0].username
                if not db.is_in_userother(tg_id):
                    print("ADD TO BASE")
                    db.create_other_user(name, tg_id, tg_name, name_group, phone)
                else:
                    print("IN BASE")
                await asyncio.sleep(1)
                await client(DeleteContactsRequest(id=[result.users[0].id]))
            except FloodWaitError as ex:
                time_sleep = int(str(ex).split("A wait of ")[1].split(" ")[0])
                print("SLEEP", time_sleep)
                await asyncio.sleep(time_sleep + 5)
                try:
                    print("GET", name, phone)
                    contact = InputPhoneContact(client_id=0, phone=phone, first_name=name.split(" ")[0],
                                                last_name=name.split(" ")[1])
                    result = await client(ImportContactsRequest([contact]))
                    tg_id = result.users[0].id
                    tg_name = result.users[0].username
                    if not db.is_in_userother(tg_id):
                        print("ADD TO BASE")
                        db.create_other_user(name, tg_id, tg_name, name_group, phone)
                    else:
                        print("IN BASE")
                    await asyncio.sleep(1)
                    await client(DeleteContactsRequest(id=[result.users[0].id]))
                except Exception as ex:
                    print(ex)
            except Exception as ex:
                print(ex)

asyncio.run(st())