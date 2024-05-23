import markups
from services.forChat.UserState import UserState
from services.forChat.Response import Response
import config_controller
from telebot import types
import services.testing.ExportDbToExel as exel
import os
from openpyxl import load_workbook
import db.database as db

class EditNameUsersState(UserState):
    async def start_msg(self):
        msg = await self.bot.send_message(chat_id=self.user_chat_id, text="Таблиця генерується...")
        doc_name = exel.create_exel("DataBase")
        await self.bot.delete_message(chat_id=self.user_chat_id, message_id=msg.id)
        return Response(documents=[open(doc_name, 'rb')], text="Змініть як вам потрібно імена у колонку G, потім завантажте їх до бота у вигляді файлу .xlsx (Exel таблиця)", buttons=markups.generate_cancel())

    async def next_msg_document(self, message: types.Message):
        await self.bot.send_message(chat_id=self.user_chat_id, text="Зміна розпочата...")
        try:
            file_name = message.document.file_name
            file_id_info = await self.bot.get_file(message.document.file_id)
            downloaded_file = await self.bot.download_file(file_id_info.file_path)
            src = file_name
            with open(src, 'wb') as new_file:
                new_file.write(downloaded_file)

            wb = load_workbook(src)
            sheet = wb["З інших чатів"]
            index = 0

            for row in sheet.iter_rows(values_only=True):
                index +=1
                if len(row) < 7:
                    break
                if row[0] == "Ім'я" or row[6] == None or row[6].startswith("="):
                    continue
                user = db.get_userother_by_name_tg(row[1])[0]
                user.name_k = row[6]
                db.session_commit(user)

            sheet = wb["Ваші контакти"]
            index = 0
            for row in sheet.iter_rows(values_only=True):
                index +=1
                if len(row) < 7:
                    break
                if row[0] == "Ім'я" or row[6] == None or row[6].startswith("="):
                    continue
                user = db.get_user_verify_by_phone(row[2])[0]
                user.name_k = row[6]
                db.session_commit(user)
            try:
                os.remove(src)
            except:
                pass
            return Response(redirect="/menu")
        except Exception as ex:
            print(ex)
            try:
                os.remove(src)
            except:
                pass
            return Response(text="Сталася помилка! Можливо ви скинули не той формат файлу", redirect="/menu")